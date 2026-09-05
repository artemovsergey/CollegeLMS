import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('import/schedule/Расписание.xlsx')
ws = wb.active

non_empty = 0
total = 0
max_weeks = 0
for row in range(6, ws.max_row+1):
    for col in range(3, ws.max_column+1):
        total += 1
        v = ws.cell(row, col).value
        if v:
            non_empty += 1

print(f'Всего ячеек: {total}, непустых: {non_empty}, пустых: {total - non_empty}')

# Показываем структуру: сколько строк на каждый (день, пара)
from collections import Counter
current_day = None
day_pairs = Counter()
for row in range(6, ws.max_row+1):
    day_val = ws.cell(row, 1).value
    pair_val = ws.cell(row, 2).value
    if day_val:
        current_day = str(day_val).strip()
    if current_day and pair_val is not None:
        day_pairs[(current_day, int(pair_val))] += 1

print(f'Уникальных (день, пара): {len(day_pairs)}')
print(f'Макс строк на (день, пара): {max(day_pairs.values())}')
for k, v in sorted(day_pairs.items()):
    if v > 1:
        print(f'  {k}: {v} строк!')
