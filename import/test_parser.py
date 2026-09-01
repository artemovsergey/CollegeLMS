import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import re

def parse_weeks(text):
    if not text:
        return []
    clean = text.strip('() ')
    weeks = []
    for part in clean.split(','):
        part = part.strip()
        if '-' in part:
            rng = part.split('-')
            try:
                for i in range(int(rng[0]), int(rng[1]) + 1):
                    weeks.append(i)
            except:
                pass
        else:
            try:
                weeks.append(int(part))
            except:
                pass
    return sorted(set(weeks))

def extract_trailing_teacher(text):
    match = re.search(r'([А-Яа-яёЁ][А-Яа-яёЁ]+\s+[А-Яа-яёЁ]\.[А-Яа-яёЁ]\.?)$', text)
    return match.group(1) if match else ''

def parse_subject_cell(cell):
    text = cell.strip()
    if not text or text == '.':
        return None

    # Handle ч.з. / с.з. / с.з
    if text.lower().startswith(('ч.з', 'с.з')):
        room = text.split()[0]
        rest = text[len(room):].strip()
        weeks_match = re.search(r'\(([^)]+)\)', rest)
        weeks = parse_weeks(weeks_match.group(1)) if weeks_match else []
        subject = re.sub(r'\([^)]*\)', '', rest).strip()
        teacher = extract_trailing_teacher(subject)
        if teacher and subject.endswith(teacher):
            subject = subject[:-len(teacher)].strip()
        return {'room': room, 'subject': subject, 'weeks': weeks, 'teacher': teacher}

    # General pattern: Room Subject(Weeks) Teacher
    match = re.match(r'^([^\s]+)\s+([^(]+?)(?:\s*\(([^)]+)\))?\s*([А-Яа-яёЁ][А-Яа-яёЁ.\s]*)?$', text)
    if match:
        room = match.group(1)
        subject = match.group(2).strip()
        weeks = parse_weeks(match.group(3)) if match.group(3) else []
        teacher = match.group(4).strip() if match.group(4) else ''
        if not teacher:
            teacher = extract_trailing_teacher(subject)
        if teacher and subject.endswith(teacher):
            subject = subject[:-len(teacher)].strip()
        return {'room': room, 'subject': subject, 'weeks': weeks, 'teacher': teacher}

    return {'room': '', 'subject': text, 'weeks': [], 'teacher': ''}


wb = openpyxl.load_workbook('import/schedule/Расписание_тест.xlsx')
ws = wb.active

# Find group columns
group_columns = {}
for col in range(3, ws.max_column + 1):
    name = ws.cell(5, col).value
    if name:
        group_columns[col] = str(name).strip()

# Find day blocks
day_map = {
    'ПОНЕДЕЛЬНИК': 'Monday', 'ВТОРНИК': 'Tuesday', 'СРЕДА': 'Wednesday',
    'ЧЕТВЕРГ': 'Thursday', 'ПЯТНИЦА': 'Friday', 'СУББОТА': 'Saturday'
}
last_row = ws.max_row
day_blocks = []
for row in range(1, last_row + 1):
    val = ws.cell(row, 1).value
    if val and str(val).strip().upper() in day_map:
        day_blocks.append((row, day_map[str(val).strip().upper()]))

# Parse
entries = []
for bi, (day_start, day) in enumerate(day_blocks):
    day_end = day_blocks[bi + 1][0] if bi + 1 < len(day_blocks) else last_row + 1

    pair_rows = []
    for r in range(day_start, day_end):
        val = ws.cell(r, 2).value
        if val is not None:
            try:
                num = int(val)
                if 1 <= num <= 7:
                    pair_rows.append(r)
            except:
                pass

    for pi in range(len(pair_rows)):
        pair_row = pair_rows[pi]
        pair_num = int(ws.cell(pair_row, 2).value)
        next_pair_row = pair_rows[pi + 1] if pi + 1 < len(pair_rows) else day_end

        for col, group_name in group_columns.items():
            for r in range(pair_row, next_pair_row):
                cell_val = ws.cell(r, col).value
                if not cell_val:
                    continue
                cell_text = str(cell_val).strip()
                if not cell_text:
                    continue

                parsed = parse_subject_cell(cell_text)
                if parsed and parsed['subject']:
                    entries.append({
                        'group': group_name,
                        'day': day,
                        'pair': pair_num,
                        'room': parsed['room'],
                        'subject': parsed['subject'],
                        'weeks': parsed['weeks'],
                        'teacher': parsed['teacher'],
                    })

# Filter for ПО 262
po262 = [e for e in entries if e['group'] == 'ПО 262']
print(f'\n=== ПО 262: {len(po262)} entries ===\n')
for e in po262:
    weeks_str = ','.join(str(w) for w in e['weeks']) if e['weeks'] else 'все'
    print(f"{e['day']:10} п{e['pair']} | {e['room']:6} | {e['subject']:25} | [{weeks_str:20}] | {e['teacher']}")

wb.close()
