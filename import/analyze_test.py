import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('import/schedule/Расписание_тест.xlsx')
ws = wb.active

# Get pair rows
pair_rows = []
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 2).value
    if v is not None and isinstance(v, (int, float)) and 1 <= v <= 7:
        pair_rows.append(r)

# For ПО 262 (column C=3), show every row with its pair context
print("=== ПО 262 (Column C) — every row ===")
for r in range(6, 176):
    cell_val = ws.cell(r, 3).value
    pair_val = ws.cell(r, 2).value
    day_val = ws.cell(r, 1).value
    
    # Find which pair this row belongs to
    current_pair = None
    current_day = None
    for pr in reversed(pair_rows):
        if pr <= r:
            current_pair = int(ws.cell(pr, 2).value)
            day_val_at_pr = ws.cell(pr, 1).value
            if day_val_at_pr:
                current_day = day_val_at_pr
            break
    
    marker = ""
    if pair_val:
        marker = f" [PAIR {int(pair_val)}]"
    if day_val:
        marker = f" [DAY: {day_val}]"
    
    cell_str = repr(cell_val)[:120] if cell_val else "empty"
    print(f"  R{r} (pair {current_pair}): {cell_str}{marker}")

wb.close()
