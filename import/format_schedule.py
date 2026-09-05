"""
Красивое форматирование расписания для печати — v2.

Улучшения:
- Современная цветовая палитра
- Чёткая иерархия шрифтов
- Оптимальные размеры ячеек
- Позиционирование: аудитория (лево), предмет (центр), недели (центр), преподаватель (право)
- Дни недели полные

Usage: python import/format_schedule.py
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re

SRC = 'import/schedule/Расписание.xlsx'
OUT = 'import/schedule/Расписание_v2.xlsx'

DAYS_RU = {
    'Monday': 'Понедельник',
    'Tuesday': 'Вторник',
    'Wednesday': 'Среда',
    'Thursday': 'Четверг',
    'Friday': 'Пятница',
}
DAY_ORDER = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
SUB_ROWS = 4

# === Цвета ===
COLORS = {
    'title_bg': '1B2A4A',      # тёмно-синий
    'header_bg': '2D4373',     # средний синий
    'day_bg': '3B6DAA',        # голубой
    'pair_bg': 'E8EDF5',       # светло-серый
    'even_bg': 'F5F7FA',       # очень светлый
    'odd_bg': 'FFFFFF',        # белый
    'title_fg': 'FFFFFF',      # белый текст
    'header_fg': 'FFFFFF',
    'day_fg': 'FFFFFF',
    'pair_fg': '1B2A4A',
    'room_fg': '7A8599',       # серый
    'subject_fg': '1A1A2E',    # почти чёрный
    'weeks_fg': '8B95A5',      # серый
    'teacher_fg': '3D5A80',    # тёмно-синий
}

# === Шрифты ===
FONTS = {
    'title': Font(name='Segoe UI', bold=True, size=14, color=COLORS['title_fg']),
    'header': Font(name='Segoe UI', bold=True, size=9, color=COLORS['header_fg']),
    'day': Font(name='Segoe UI', bold=True, size=10, color=COLORS['day_fg']),
    'pair': Font(name='Segoe UI', bold=True, size=10, color=COLORS['pair_fg']),
    'room': Font(name='Segoe UI', size=7, color=COLORS['room_fg']),
    'subject': Font(name='Segoe UI', bold=True, size=9, color=COLORS['subject_fg']),
    'weeks': Font(name='Segoe UI', size=7, color=COLORS['weeks_fg'], italic=True),
    'teacher': Font(name='Segoe UI', size=8, color=COLORS['teacher_fg']),
}

# === Заливки ===
FILLS = {
    'title': PatternFill(start_color=COLORS['title_bg'], end_color=COLORS['title_bg'], fill_type='solid'),
    'header': PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type='solid'),
    'day': PatternFill(start_color=COLORS['day_bg'], end_color=COLORS['day_bg'], fill_type='solid'),
    'pair': PatternFill(start_color=COLORS['pair_bg'], end_color=COLORS['pair_bg'], fill_type='solid'),
    'even': PatternFill(start_color=COLORS['even_bg'], end_color=COLORS['even_bg'], fill_type='solid'),
    'odd': PatternFill(start_color=COLORS['odd_bg'], end_color=COLORS['odd_bg'], fill_type='solid'),
}

# === Границы ===
THIN = Side(style='thin', color='D0D5DD')
HAIR = Side(style='hair', color='E4E7EC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_LIGHT = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)


def parse_cell(value: str) -> dict:
    if not value or not isinstance(value, str):
        return {'room': '', 'subject': '', 'weeks': '', 'teacher': ''}
    v = value.strip()
    if not v:
        return {'room': '', 'subject': '', 'weeks': '', 'teacher': ''}

    weeks_match = re.search(r'\(([^)]+)\)', v)
    weeks = ''
    before_weeks = v
    after_weeks = ''
    if weeks_match:
        weeks = f'({weeks_match.group(1)})'
        before_weeks = v[:weeks_match.start()].strip()
        after_weeks = v[weeks_match.end():].strip()

    before_parts = before_weeks.split()
    room = before_parts[0] if before_parts else ''
    subject = ' '.join(before_parts[1:]) if len(before_parts) > 1 else ''
    teacher = after_weeks.strip()

    return {'room': room, 'subject': subject, 'weeks': weeks, 'teacher': teacher}


def create_schedule():
    wb_in = openpyxl.load_workbook(SRC)
    ws_in = wb_in.active

    groups = []
    for col in range(3, ws_in.max_column + 1):
        val = ws_in.cell(5, col).value
        if val:
            groups.append((col, str(val).strip()))

    schedule_data = {}
    current_day = None
    for row in range(6, ws_in.max_row + 1):
        day_val = ws_in.cell(row, 1).value
        pair_val = ws_in.cell(row, 2).value
        if day_val:
            day_str = str(day_val).strip().upper()
            for key, ru in DAYS_RU.items():
                if ru.upper() in day_str:
                    current_day = key
                    break
        if current_day and pair_val is not None:
            pair = int(pair_val)
            schedule_data[(current_day, pair)] = {}
            for col, _ in groups:
                cell_val = ws_in.cell(row, col).value
                if cell_val:
                    schedule_data[(current_day, pair)][col] = str(cell_val).strip()

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Расписание'

    ws_out.page_setup.orientation = 'landscape'
    ws_out.page_setup.paperSize = ws_out.PAPERSIZE_A3
    ws_out.page_setup.fitToWidth = 1
    ws_out.page_setup.fitToHeight = 0
    ws_out.sheet_properties.pageSetUpPr.fitToPage = True
    ws_out.page_margins.left = 0.2
    ws_out.page_margins.right = 0.2
    ws_out.page_margins.top = 0.3
    ws_out.page_margins.bottom = 0.3

    num_cols = len(groups)
    out_col_count = num_cols + 1

    # === Заголовок ===
    r = 1
    ws_out.merge_cells(start_row=r, start_column=1, end_row=r, end_column=out_col_count)
    c = ws_out.cell(r, 1, 'РАСПИСАНИЕ УЧЕБНЫХ ЗАНЯТИЙ')
    c.font = FONTS['title']
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = FILLS['title']
    ws_out.row_dimensions[r].height = 30
    r += 1

    # === Шапка групп ===
    ws_out.cell(r, 1, '№').font = FONTS['header']
    ws_out.cell(r, 1).fill = FILLS['header']
    ws_out.cell(r, 1).border = BORDER
    ws_out.cell(r, 1).alignment = Alignment(horizontal='center', vertical='center')
    for i, (_, gname) in enumerate(groups):
        c = ws_out.cell(r, i + 2, gname)
        c.font = FONTS['header']
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.fill = FILLS['header']
        c.border = BORDER
    ws_out.row_dimensions[r].height = 20
    r += 1

    max_pair = max((p for _, p in schedule_data.keys()), default=7)

    for day_key in DAY_ORDER:
        day_ru = DAYS_RU[day_key]

        # Строка дня
        ws_out.merge_cells(start_row=r, start_column=1, end_row=r, end_column=out_col_count)
        c = ws_out.cell(r, 1, day_ru.upper())
        c.font = FONTS['day']
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = FILLS['day']
        c.border = BORDER
        ws_out.row_dimensions[r].height = 22
        r += 1

        for pair_num in range(1, max_pair + 1):
            pair_data = schedule_data.get((day_key, pair_num), {})
            fill = FILLS['even'] if pair_num % 2 == 0 else FILLS['odd']

            for sub in range(SUB_ROWS):
                if sub == 0:
                    ws_out.merge_cells(
                        start_row=r, start_column=1,
                        end_row=r + SUB_ROWS - 1, end_column=1
                    )
                    c = ws_out.cell(r, 1, str(pair_num))
                    c.font = FONTS['pair']
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.fill = FILLS['pair']
                    c.border = BORDER

                for i, (col, _) in enumerate(groups):
                    out_c = i + 2
                    raw = pair_data.get(col, '')
                    parsed = parse_cell(raw) if raw else {
                        'room': '', 'subject': '', 'weeks': '', 'teacher': ''
                    }

                    if sub == 0:
                        text = parsed['room']
                        font = FONTS['room']
                        halign = 'left'
                        valign = 'top'
                    elif sub == 1:
                        text = parsed['subject']
                        font = FONTS['subject']
                        halign = 'center'
                        valign = 'center'
                    elif sub == 2:
                        text = parsed['weeks']
                        font = FONTS['weeks']
                        halign = 'center'
                        valign = 'center'
                    else:
                        text = parsed['teacher']
                        font = FONTS['teacher']
                        halign = 'right'
                        valign = 'bottom'

                    c = ws_out.cell(r, out_c, text)
                    c.font = font
                    c.alignment = Alignment(
                        horizontal=halign, vertical=valign, wrap_text=True
                    )
                    c.fill = fill
                    c.border = BORDER_LIGHT

                heights = {0: 12, 1: 26, 2: 12, 3: 14}
                ws_out.row_dimensions[r].height = heights[sub]
                r += 1

    ws_out.column_dimensions[get_column_letter(1)].width = 3.5
    for i in range(num_cols):
        ws_out.column_dimensions[get_column_letter(i + 2)].width = 15

    ws_out.oddHeader.center.text = 'РАСПИСАНИЕ УЧЕБНЫХ ЗАНЯТИЙ'
    ws_out.oddFooter.center.text = 'Стр. &P из &N'

    wb_out.save(OUT)
    print(f'Готово: {OUT}')
    print(f'Групп: {num_cols}, Дней: {len(DAY_ORDER)}, Пар: {max_pair}')
    print(f'Строк: {r}, Столбцов: {out_col_count}')


if __name__ == '__main__':
    create_schedule()
