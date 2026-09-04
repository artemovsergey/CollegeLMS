"""
Fix formatting issues in Расписание.xlsx.

Issues fixed:
1. с.з without dot → с.з.
2. Latin c in c.з. → Cyrillic с → с.з.
3. Bare Физ.кул./Физкультура without с.з. prefix → add с.з.
4. Missing space before ( where attached to subject name
5. Dot in weeks like (4.) → (4)
6. Double spaces → single space
7. Trailing space+dot after teacher initials → no space

Usage: python import/fix_schedule_format.py
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import re
import shutil
from datetime import datetime

SRC = 'import/schedule/Расписание.xlsx'

def fix_cell(value: str) -> str:
    v = value.strip()

    # 1. Latin c → Cyrillic с in c.з.
    v = v.replace('c.з.', 'с.з.')

    # 2. с.з without dot → с.з.
    v = re.sub(r'с\.з(?!\.)', 'с.з.', v)

    # 3. Bare Физ.кул./Физкультура without с.з. prefix
    # Only if not already prefixed with с.з.
    if 'с.з' not in v:
        # Физ.кул. (weeks) teacher → с.з. Физ.кул. (weeks) teacher
        v = re.sub(r'^(Физ\.кул\.|Физ\.культура)\s', r'с.з. \1 ', v)

    # 4. Missing space before ( where attached to subject name
    # e.g. "ОБПиЗР(13)" → "ОБПиЗР (13)", "История(1)" → "История (1)"
    # Also "Матем.(3,5,7)" → "Матем. (3,5,7)" (subject abbreviation ending with .)
    # Also "МДК.05.01(12)" → "МДК.05.01 (12)" (digit before ( without space)
    # But NOT "МДК.05.01.(2,4,8,9)" — that's room.code.(weeks) which is fine
    v = re.sub(r'([а-яА-Я])\(', r'\1 (', v)
    v = re.sub(r'([а-яА-Я])\.(\()', r'\1. \2', v)
    v = re.sub(r'(\d)\(', r'\1 (', v)

    # 7b. Missing period after teacher initials: "Кривцова С.Н" → "Кривцова С.Н."
    v = re.sub(r'([А-Я]\.[А-Я])(?=\s*$)', r'\1.', v)

    # 5. Dot in weeks like (4.) → (4), (10.) → (10), (17.) → (17)
    v = re.sub(r'\((\d+)\.\)', r'(\1)', v)

    # 6. Double spaces → single space
    v = re.sub(r'  +', ' ', v)

    # 7. Trailing space+dot after teacher initials: "В.Ф ." → "В.Ф."
    v = re.sub(r'([А-Я]\.[А-Я])\s+\.\s*$', r'\1.', v)

    return v


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb.active

    fixed_count = 0
    changes_by_type = {
        'latin_c': 0,
        'missing_dot_cz': 0,
        'bare_fizkultura': 0,
        'missing_space_paren': 0,
        'dot_in_weeks': 0,
        'double_space': 0,
        'trailing_space_dot': 0,
    }

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not v or not isinstance(v, str):
                continue

            original = v.strip()
            fixed = fix_cell(original)

            if fixed != original:
                # Track what changed
                if 'c.з.' in original and 'c.з.' not in fixed:
                    changes_by_type['latin_c'] += 1
                if re.search(r'с\.з(?!\.)', original):
                    changes_by_type['missing_dot_cz'] += 1
                if re.search(r'^Физ\.', original) and 'с.з' not in original:
                    changes_by_type['bare_fizkultura'] += 1
                if re.search(r'[а-яА-Я]\(', original):
                    changes_by_type['missing_space_paren'] += 1
                if re.search(r'\(\d+\.\)', original):
                    changes_by_type['dot_in_weeks'] += 1
                if '  ' in original:
                    changes_by_type['double_space'] += 1
                if re.search(r'[А-Я]\.[А-Я]\.\s+\.', original):
                    changes_by_type['trailing_space_dot'] += 1

                ws.cell(r, c).value = fixed
                fixed_count += 1

    # Save with backup
    backup = SRC.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    shutil.copy2(SRC, backup)
    print(f'Backup saved: {backup}')

    wb.save(SRC)
    print(f'Fixed {fixed_count} cells in {SRC}')
    print()
    print('Changes by type:')
    for name, count in changes_by_type.items():
        if count > 0:
            print(f'  {name}: {count}')


if __name__ == '__main__':
    main()
