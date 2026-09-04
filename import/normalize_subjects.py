"""
Normalize subject names in Расписание.xlsx.

Rules:
- Физ.кул. → Физкультура
- Физ.культура → Физкультура
- с.з. Физ.кул. → с.з. Физкультура
- с.з. Физ.культура → с.з. Физкультура
- Ин.язык. → Ин.язык
- Ист.Р. → История
- с.з.Физ.культура → с.з. Физкультура (fix spacing)

Usage: python import/normalize_subjects.py
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import re

SRC = 'import/schedule/Расписание.xlsx'

def normalize(value: str) -> str:
    v = value.strip()

    # Fix Физкультура variations
    # с.з. Физ.кул. → с.з. Физкультура
    v = re.sub(r'(с\.з\.\s*)Физ\.кул\.', r'\1Физкультура', v)
    # с.з. Физ.культура → с.з. Физкультура
    v = re.sub(r'(с\.з\.\s*)Физ\.культура', r'\1Физкультура', v)
    # Bare Физ.кул. → Физкультура (without с.з. prefix)
    v = re.sub(r'(?<!с\.з\.)Физ\.кул\.', 'Физкультура', v)
    # Bare Физ.культура → Физкультура
    v = re.sub(r'(?<!с\.з\.)Физ\.культура', 'Физкультура', v)
    # с.з.Физ.культура (no space) → с.з. Физкультура
    v = re.sub(r'с\.з\.\s*Физ', 'с.з. Физ', v)

    # Fix Ин.язык variations
    v = re.sub(r'Ин\.язык\.', 'Ин.язык', v)

    # Fix История variations
    v = re.sub(r'Ист\.Р\.', 'ИсторияРоссии', v)

    # Fix Математика variations
    v = re.sub(r'Матем\.(?!\d)', 'Математика', v)

    # Fix ОхранаТруда variations
    v = re.sub(r'Охр\.тр\.', 'ОхранаТруда', v)
    v = re.sub(r'Охрана труда', 'ОхранаТруда', v)

    # Fix ОсновыЭлектр. variations
    v = re.sub(r'ОсновыЭлект\.', 'ОсновыЭлектр.', v)

    # Fix ЭкономОтр. variations
    v = re.sub(r'Эконом\. отр\.', 'ЭкономОтр.', v)
    v = re.sub(r'Эконом\.отр\.', 'ЭкономОтр.', v)

    # Fix ЭлектрТех. variations (order matters: longer patterns first)
    v = re.sub(r'ЭлектротехиЭ\.', 'ЭлектрТех.', v)
    v = re.sub(r'Электр\.и Э\.', 'ЭлектрТех.', v)
    v = re.sub(r'Электротех\.(?!и)', 'ЭлектрТех.', v)
    v = re.sub(r'Электр\.тех\.', 'ЭлектрТех.', v)
    v = re.sub(r'Электр\.(?!\d|Т|т)', 'ЭлектрТех.', v)
    v = re.sub(r'Эл\.тех\.', 'ЭлектрТех.', v)
    v = re.sub(r'Электробез\.', 'ЭлектрБезопасность', v)

    # Fix ОсновыЭлектрТех. (was incorrectly changed from ОсновыЭлектр.)
    v = re.sub(r'ОсновыЭлектрТех\.', 'ОсновыЭлектр.', v)

    return v


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb.active

    fixed = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not v or not isinstance(v, str):
                continue
            new_v = normalize(v)
            if new_v != v:
                ws.cell(r, c).value = new_v
                fixed += 1

    wb.save(SRC)
    print('Normalized %d cells in %s' % (fixed, SRC))
    wb.close()


if __name__ == '__main__':
    main()
