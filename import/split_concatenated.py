import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import re
from copy import copy
from openpyxl.utils import get_column_letter

def split_subjects(cell_text):
    """Split concatenated multi-subject cell text into individual subjects."""
    if not cell_text:
        return [None]
    text = cell_text.strip()
    starts = [0]
    # Room number at word boundary
    for m in re.finditer(r'(?<!\d)(?<!,)\s(\d{3}[лМм]?\s)', text):
        starts.append(m.start(1))
    # 'с.з.' not at position 0
    for m in re.finditer(r'(?<!^)\s*(с\.з\.)\s', text):
        if m.start(1) > 0:
            starts.append(m.start(1))
    # Subject names without room number
    for m in re.finditer(r'(?<=[а-яА-Я.])\s+(Физ\.культура|Физ\.кул\.|Физкультура|Ин\.язык)[\s(]', text):
        starts.append(m.start(1))
    # Teacher with initials (weeks) teacher with initials pattern — split before weeks
    for m in re.finditer(r'([А-Я][а-я]+\s+[А-Я]\.[А-Я]\.?)\s+(\(\d[\d,\-]*\))\s+([А-Я][а-я]+\s+[А-Я]\.[А-Я]\.?)', text):
        if m.start(2) > 0:
            starts.append(m.start(2))
    
    starts = sorted(set(starts))
    if len(starts) <= 1:
        # No split points found, try special case
        return _split_special_case(text)
    
    subjects = []
    for i, start in enumerate(starts):
        end = starts[i+1] if i+1 < len(starts) else len(text)
        part = text[start:end].strip()
        if part:
            # Recursively split each part if it's still multi-subject
            sub_parts = _split_special_case(part) if is_multi_subject(part) else [part]
            subjects.extend(sub_parts)
    return subjects


def _split_special_case(text):
    """Handle case: room subject1 (weeks1) subject2 (weeks2) teacher — same room, one teacher."""
    room_matches = list(re.finditer(r'\d{3}[лМм]?\s', text))
    if len(room_matches) == 1:
        weeks_parts = re.split(r'(\(\d[\d,\-]*\))', text)
        if len(weeks_parts) >= 5:
            teacher_match = re.search(r'([А-Я][а-я]+\s+[А-Я]\.[А-Я]\.?)$', text)
            if teacher_match:
                teacher = teacher_match.group(1)
                subj1 = (weeks_parts[0] + weeks_parts[1] + ' ' + teacher).strip()
                subj2 = (weeks_parts[2] + weeks_parts[3] + ' ' + teacher).strip()
                return [subj1, subj2]
    return [text]

def is_multi_subject(cell_text):
    if not cell_text:
        return False
    weeks_pattern = r'\(\d[\d,\-]*\)\s?'
    weeks = re.findall(weeks_pattern, cell_text)
    return len(weeks) >= 2

def copy_style(src, tgt):
    if src.has_style:
        tgt.font = copy(src.font)
        tgt.border = copy(src.border)
        tgt.fill = copy(src.fill)
        tgt.number_format = copy(src.number_format)
        tgt.protection = copy(src.protection)
        tgt.alignment = copy(src.alignment)

def main():
    input_file = 'import/schedule/Расписание_все_группы.xlsx'
    output_file = 'import/schedule/Расписание_split.xlsx'

    print(f'Reading {input_file}...')
    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in.active

    orig_max_row = ws_in.max_row
    orig_max_col = ws_in.max_column
    print(f'Input: {orig_max_row} rows x {orig_max_col} cols')

    # Collect all merged cell ranges from input
    orig_merges = list(ws_in.merged_cells.ranges)

    # Collect column widths
    col_widths = {}
    for c in range(1, orig_max_col + 1):
        letter = get_column_letter(c)
        if letter in ws_in.column_dimensions:
            col_widths[c] = ws_in.column_dimensions[letter].width

    # Collect row heights
    row_heights = {}
    for r in range(1, orig_max_row + 1):
        if r in ws_in.row_dimensions:
            row_heights[r] = ws_in.row_dimensions[r].height

    # Step 1: For each data row (6..max), compute how many output rows it needs
    # row_expansion[orig_row] = number of output rows for this orig row (1 if no split, N if N subjects)
    row_expansion = {}
    # Also collect per-cell splits: {(orig_row, col): [subject1, subject2, ...]}
    cell_splits = {}

    for r in range(1, orig_max_row + 1):
        max_subjects = 1
        for c in range(3, orig_max_col + 1):
            v = ws_in.cell(r, c).value
            if v and isinstance(v, str) and is_multi_subject(v):
                subjects = split_subjects(v)
                cell_splits[(r, c)] = subjects
                max_subjects = max(max_subjects, len(subjects))
        row_expansion[r] = max_subjects

    # Step 2: Build mapping from orig_row -> output_row start
    row_map = {}  # orig_row -> first output_row
    out_row = 1
    for r in range(1, orig_max_row + 1):
        row_map[r] = out_row
        out_row += row_expansion[r]

    total_out_rows = out_row - 1
    print(f'Output will have {total_out_rows} rows (inserted {total_out_rows - orig_max_row} extra)')

    # Step 3: Create output workbook
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_in.title

    # Copy column widths
    for c, w in col_widths.items():
        letter = get_column_letter(c)
        ws_out.column_dimensions[letter].width = w

    # Step 4: Copy rows
    for r in range(1, orig_max_row + 1):
        out_start = row_map[r]
        expansion = row_expansion[r]

        for sub_idx in range(expansion):
            out_r = out_start + sub_idx

            # Copy row height (only for first sub-row)
            if sub_idx == 0 and r in row_heights:
                ws_out.row_dimensions[out_r].height = row_heights[r]

            for c in range(1, orig_max_col + 1):
                src_cell = ws_in.cell(r, c)

                # Determine value
                if (r, c) in cell_splits:
                    subjects = cell_splits[(r, c)]
                    if sub_idx < len(subjects):
                        val = subjects[sub_idx]
                    else:
                        val = None
                else:
                    # For merged cells that extend across multiple orig rows:
                    # only write value on the first output row
                    if sub_idx == 0:
                        val = src_cell.value
                    else:
                        val = None

                tgt_cell = ws_out.cell(out_r, c, val)
                copy_style(src_cell, tgt_cell)

    # Step 5: Re-create merged cells
    for merge in orig_merges:
        min_r = merge.min_row
        max_r = merge.max_row
        col = merge.min_col

        out_min = row_map[min_r]
        # For merged cells spanning multiple rows, compute the expanded end
        out_max = row_map[min_r]
        for mr in range(min_r, max_r + 1):
            out_max = row_map[mr] + row_expansion[mr] - 1

        if out_min == out_max and merge.min_col == merge.max_col:
            # Single cell, no merge needed
            pass
        else:
            try:
                ws_out.merge_cells(
                    start_row=out_min,
                    start_column=merge.min_col,
                    end_row=out_max,
                    end_column=merge.max_col
                )
                # Copy style from first cell
                src = ws_in.cell(min_r, merge.min_col)
                tgt = ws_out.cell(out_min, merge.min_col)
                copy_style(src, tgt)
            except Exception as e:
                print(f'  Warning: could not merge {merge}: {e}')

    # Step 6: Verify
    multi_count = 0
    single_count = 0
    for c in range(3, ws_out.max_column + 1):
        for r in range(6, ws_out.max_row + 1):
            v = ws_out.cell(r, c).value
            if v and isinstance(v, str) and v.strip():
                if is_multi_subject(v):
                    multi_count += 1
                else:
                    single_count += 1

    print(f'Output: {ws_out.max_row} rows x {ws_out.max_column} cols')
    print(f'After split: {single_count} single-subject, {multi_count} multi-subject')

    print(f'Saving to {output_file}...')
    wb_out.save(output_file)
    wb_out.close()
    wb_in.close()
    print('Done!')

if __name__ == '__main__':
    main()
